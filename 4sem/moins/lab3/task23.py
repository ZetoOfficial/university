import csv

from openpyxl import load_workbook

N = len('Павел')
M = len('Титов')

wb = load_workbook('out.xlsx')
ws = wb.get_sheet_by_name("Sheet")

table = []
for x in range(1, N + 1):
    table.append([])
    for y in range(1, M + 1):
        table[x - 1].append(int(ws.cell(row=x, column=y).value))

with open('out.csv', 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerows(table)
